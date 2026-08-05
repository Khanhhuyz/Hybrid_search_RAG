"""Safe local artifact writers and validators."""

import csv
import re
import uuid
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from app.config import settings


def safe_name(value: str) -> str:
    name = re.sub(r"[^a-zA-Z0-9._-]+", "-", value.strip()).strip("-.")[:80]
    return name or "artifact"


def markdown_content(data: dict[str, Any], artifact_type: str) -> str:
    lines = [f"# {data.get('title', 'Agent Artifact')}", "", data.get("summary", ""), ""]
    table = data.get("table") or {}
    headers, rows = table.get("headers") or [], table.get("rows") or []
    if headers:
        lines += [
            "| " + " | ".join(map(str, headers)) + " |",
            "| " + " | ".join(["---"] * len(headers)) + " |",
        ]
        lines += [
            "| " + " | ".join(str(cell).replace("|", "\\|") for cell in row) + " |"
            for row in rows
        ]
        lines.append("")
    for section in data.get("sections") or []:
        lines += [f"## {section.get('heading', 'Section')}", "", section.get("content", ""), ""]
    plan = data.get("plan") or {}
    if plan:
        lines += ["## Goal", "", plan.get("goal", ""), "", "## Milestones", ""]
        for item in plan.get("milestones") or []:
            lines += [f"### {item.get('name', 'Milestone')} — {item.get('deadline', 'TBD')}"]
            lines += [f"- {task}" for task in item.get("tasks") or []]
            lines.append("")
    citations = data.get("citations") or []
    if citations:
        lines += ["## Sources", ""] + [
            f"- [S{i + 1}] {item}" for i, item in enumerate(citations)
        ]
    return "\n".join(lines).strip() + "\n"


class ArtifactWriter:
    def __init__(self):
        self.root = settings.OUTPUT_DIR.resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def write(
        self,
        run_id: str,
        artifact_type: str,
        data: dict[str, Any],
        output_format: str,
    ) -> dict:
        artifact_id = str(uuid.uuid4())
        title = str(data.get("title") or "Agent Artifact")
        base = f"{safe_name(title)}-{artifact_id[:8]}"
        creators = {
            "md": self._markdown,
            "csv": self._csv,
            "xlsx": self._xlsx,
            "pdf": self._pdf,
            "svg": self._svg,
        }
        if output_format not in creators:
            raise ValueError(f"Unsupported artifact format: {output_format}")
        path, mime = creators[output_format](base, data, artifact_type)
        resolved = path.resolve()
        if self.root not in resolved.parents or not resolved.is_file():
            raise ValueError("Artifact path escaped the output directory")
        size = resolved.stat().st_size
        if size <= 0 or size > settings.MAX_OUTPUT_FILE_SIZE_MB * 1024 * 1024:
            raise ValueError("Artifact file failed size validation")
        if output_format == "xlsx":
            load_workbook(resolved, read_only=True).close()
        return {
            "id": artifact_id,
            "run_id": run_id,
            "artifact_type": artifact_type,
            "title": title,
            "filename": resolved.name,
            "file_path": str(resolved),
            "mime_type": mime,
            "size": size,
            "preview": markdown_content(data, artifact_type)[:5000],
        }

    def _markdown(self, base, data, artifact_type):
        path = self.root / f"{base}.md"
        path.write_text(markdown_content(data, artifact_type), encoding="utf-8")
        return path, "text/markdown; charset=utf-8"

    def _csv(self, base, data, artifact_type):
        table = data.get("table") or {}
        path = self.root / f"{base}.csv"
        with path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(table.get("headers") or ["Content"])
            rows = table.get("rows") or [[data.get("summary", "")]]
            writer.writerows(rows)
        return path, "text/csv; charset=utf-8"

    def _xlsx(self, base, data, artifact_type):
        path = self.root / f"{base}.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Artifact"
        table = data.get("table") or {}
        headers = table.get("headers") or ["Content"]
        rows = table.get("rows") or [[data.get("summary", "")]]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
        for row in rows:
            worksheet.append(list(row))
        for column in worksheet.columns:
            worksheet.column_dimensions[column[0].column_letter].width = min(
                60,
                max(12, max(len(str(cell.value or "")) for cell in column) + 2),
            )
        workbook.save(path)
        return path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _pdf(self, base, data, artifact_type):
        path = self.root / f"{base}.pdf"
        styles, story = getSampleStyleSheet(), []
        font_candidates = [
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("C:/Windows/Fonts/arial.ttf"),
            Path("C:/Windows/Fonts/calibri.ttf"),
        ]
        font_path = next((candidate for candidate in font_candidates if candidate.exists()), None)
        if font_path:
            font_name = "AgentUnicode"
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
            for style_name in ("Heading1", "Heading2", "BodyText"):
                styles[style_name].fontName = font_name
        content = markdown_content(data, artifact_type)
        for line in content.splitlines():
            clean = escape(line.lstrip("# ") or " ")
            style = (
                styles["Heading1"]
                if line.startswith("# ")
                else styles["Heading2"]
                if line.startswith("## ")
                else styles["BodyText"]
            )
            story.extend([Paragraph(clean, style), Spacer(1, 6)])
        SimpleDocTemplate(
            str(path), pagesize=A4, title=str(data.get("title", "Artifact"))
        ).build(story)
        return path, "application/pdf"

    def _svg(self, base, data, artifact_type):
        path = self.root / f"{base}.svg"
        table = data.get("table") or {}
        rows = table.get("rows") or []
        values = []
        for index, row in enumerate(rows[:12]):
            number = next(
                (float(value) for value in row if isinstance(value, (int, float))),
                float(index + 1),
            )
            values.append((str(row[0])[:24] if row else str(index + 1), number))
        maximum = max([value for _, value in values] or [1])
        bars = "".join(
            f'<text x="10" y="{55 + index * 32}" fill="#ddd">{escape(label)}</text>'
            f'<rect x="190" y="{38 + index * 32}" width="{max(2, value / maximum * 500):.1f}" '
            f'height="20" fill="#6366f1"/>'
            for index, (label, value) in enumerate(values)
        )
        svg = (
            f'<svg xmlns="http://www.w3.org/2000/svg" width="760" '
            f'height="{max(120, 80 + len(values) * 32)}" style="background:#111318">'
            f'<text x="10" y="25" fill="white" font-size="18">'
            f'{escape(str(data.get("title", "Chart")))}</text>{bars}</svg>'
        )
        path.write_text(svg, encoding="utf-8")
        return path, "image/svg+xml"
